import re
import openpyxl
from openpyxl.chart import LineChart, Reference, BarChart

class Excel:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.sh = self.wb.active
        self.sh.title = 'LandCal'
        self.sh.append(["매도가격", "수익"])

    def saveData(self, data):
        self.sh = self.wb.active
        self.sh.title = 'LandCal'
        self.sh.append(["매도가격", "수익"])

        for d in data:
            for i in range(len(d)):
                self.sh.append([d[i][0], d[i][1]])

    def gridChart(self):
        chart = LineChart()
        cData = Reference(self.sh, min_col=2, min_row=1, max_row=self.sh.max_row)
        chart.add_data(cData, titles_from_data=True)
        cat1 = Reference(self.sh, min_row=2, max_row=self.sh.max_row, min_col=1)
        chart.set_categories(cat1)
        self.sh.add_chart(chart, 'C1')

    def write(self):
        self.wb.save('LandCal.xlsx')



